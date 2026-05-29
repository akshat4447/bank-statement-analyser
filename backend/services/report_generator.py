"""
Report generator: PDF (Perfios-style) and Excel export.
"""
import os
from datetime import datetime
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from models.schemas import AnalysisResponse

# Color palette
NAVY = colors.HexColor("#1a2e4a")
TEAL = colors.HexColor("#0d7377")
LIGHT_BLUE = colors.HexColor("#e8f4f8")
GREEN = colors.HexColor("#27ae60")
RED = colors.HexColor("#e74c3c")
AMBER = colors.HexColor("#f39c12")
GRAY = colors.HexColor("#95a5a6")
LIGHT_GRAY = colors.HexColor("#f5f6fa")


def _risk_color(risk: str):
    return {
        "LOW": GREEN,
        "MEDIUM": AMBER,
        "HIGH": RED,
        "VERY HIGH": RED,
    }.get(risk, GRAY)


def generate_pdf_report(analysis: AnalysisResponse, output_path: str) -> str:
    """Generate a Perfios-style PDF report."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", fontSize=18, textColor=NAVY, spaceAfter=6, fontName="Helvetica-Bold")
    h1_style = ParagraphStyle("H1", fontSize=13, textColor=NAVY, spaceAfter=4, fontName="Helvetica-Bold")
    h2_style = ParagraphStyle("H2", fontSize=11, textColor=TEAL, spaceAfter=4, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body", fontSize=9, spaceAfter=3, fontName="Helvetica")
    small_style = ParagraphStyle("Small", fontSize=8, textColor=GRAY, fontName="Helvetica")

    story = []
    ai = analysis.account_info
    an = analysis.analytics
    qa = analysis.qa_result
    cw = an.creditworthiness if an else None

    # Header
    story.append(Paragraph("Bank Statement Analysis Report", title_style))
    story.append(Paragraph("AI-Powered Financial Intelligence Platform", small_style))
    story.append(HRFlowable(width="100%", thickness=2, color=NAVY, spaceAfter=12))

    # Account summary table
    if ai:
        story.append(Paragraph("Account Information", h1_style))
        acct_data = [
            ["Account Holder", ai.account_holder or "—", "Bank", ai.bank_name or "—"],
            ["Account Number", ai.account_number or "—", "Account Type", ai.account_type or "—"],
            ["Branch", ai.branch or "—", "IFSC", ai.ifsc or "—"],
            ["Statement From", ai.statement_period_from or "—", "Statement To", ai.statement_period_to or "—"],
            ["Opening Balance", f"₹{ai.opening_balance:,.2f}" if ai.opening_balance else "—",
             "Closing Balance", f"₹{ai.closing_balance:,.2f}" if ai.closing_balance else "—"],
        ]
        acct_table = Table(acct_data, colWidths=[3.5 * cm, 5 * cm, 3.5 * cm, 5 * cm])
        acct_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_BLUE),
            ("BACKGROUND", (2, 0), (2, -1), LIGHT_BLUE),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(acct_table)
        story.append(Spacer(1, 12))

    # Key financial metrics
    if an:
        story.append(Paragraph("Financial Summary", h1_style))
        metrics_data = [
            ["Metric", "Value", "Metric", "Value"],
            ["Total Credits", f"₹{an.total_credits:,.2f}", "Total Debits", f"₹{an.total_debits:,.2f}"],
            ["Net Cash Flow", f"₹{an.net_cash_flow:,.2f}", "Avg Monthly Balance", f"₹{an.average_monthly_balance:,.2f}"],
            ["Min Balance", f"₹{an.min_balance:,.2f}", "Max Balance", f"₹{an.max_balance:,.2f}"],
            ["Total Transactions", str(an.total_transactions), "Analysis Period", f"{an.analysis_period_months} month(s)"],
            ["Bounce Count", str(len(an.bounce_transactions)), "Suspicious Transactions", str(len(an.suspicious_transactions))],
        ]
        m_table = Table(metrics_data, colWidths=[4.5 * cm, 4.5 * cm, 4.5 * cm, 4.5 * cm])
        m_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (0, -1), LIGHT_BLUE),
            ("BACKGROUND", (2, 1), (2, -1), LIGHT_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ]))
        story.append(m_table)
        story.append(Spacer(1, 12))

    # Creditworthiness section
    if cw:
        story.append(Paragraph("Creditworthiness & Risk Assessment", h1_style))

        risk_color = _risk_color(cw.risk_category)

        cw_data = [
            ["BSA Score", f"{cw.bsa_score}/100", "Risk Category", cw.risk_category],
            ["FOIR", f"{cw.foir}%" if cw.foir else "N/A", "Income Stability", f"{cw.income_stability_index}/100"],
            ["Avg Monthly Income", f"₹{cw.average_monthly_income:,.2f}", "Avg Monthly EMI", f"₹{cw.average_monthly_emi:,.2f}"],
            ["Disposable Income", f"₹{cw.disposable_income:,.2f}", "Max Eligible EMI", f"₹{cw.max_eligible_emi:,.2f}"],
        ]
        cw_table = Table(cw_data, colWidths=[4.5 * cm, 4.5 * cm, 4.5 * cm, 4.5 * cm])
        cw_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_BLUE),
            ("BACKGROUND", (2, 0), (2, -1), LIGHT_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(cw_table)
        story.append(Spacer(1, 6))

        if cw.risk_flags:
            story.append(Paragraph("Risk Flags", h2_style))
            for flag in cw.risk_flags:
                sev_color = RED if flag.severity == "high" else AMBER if flag.severity == "medium" else GRAY
                flag_data = [[flag.flag_type, flag.description, flag.severity.upper()]]
                flag_table = Table(flag_data, colWidths=[5 * cm, 10 * cm, 2.5 * cm])
                flag_table.setStyle(TableStyle([
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("TEXTCOLOR", (2, 0), (2, -1), sev_color),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("PADDING", (0, 0), (-1, -1), 4),
                    ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GRAY),
                ]))
                story.append(flag_table)
                story.append(Spacer(1, 2))
        story.append(Spacer(1, 12))

    # Monthly summary
    if an and an.monthly_stats:
        story.append(Paragraph("Monthly Summary", h1_style))
        month_header = ["Month", "Credits (₹)", "Debits (₹)", "Net Flow (₹)", "Avg Balance (₹)", "Salary (₹)", "EMI (₹)", "Bounces"]
        month_data = [month_header]
        for ms in an.monthly_stats:
            month_data.append([
                ms.month,
                f"{ms.total_credits:,.0f}",
                f"{ms.total_debits:,.0f}",
                f"{ms.net_cash_flow:,.0f}",
                f"{ms.average_balance:,.0f}",
                f"{ms.salary_credits:,.0f}",
                f"{ms.emi_debits:,.0f}",
                str(ms.bounce_count),
            ])
        col_w = [2.5, 2.2, 2.2, 2.2, 2.5, 2.2, 2.0, 1.5]
        m_table = Table(month_data, colWidths=[w * cm for w in col_w])
        m_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ("PADDING", (0, 0), (-1, -1), 3),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(m_table)
        story.append(Spacer(1, 12))

    # QA Validation
    if qa:
        story.append(PageBreak())
        story.append(Paragraph("AI QA Validation Report", h1_style))

        qa_summary = [
            ["Overall Confidence", f"{qa.overall_confidence}%", "Data Quality Grade", qa.data_quality_grade],
            ["Extraction Accuracy", f"{qa.extraction_accuracy}%", "Calculation Accuracy", f"{qa.calculation_accuracy}%"],
            ["Categorization Accuracy", f"{qa.categorization_accuracy}%", "Issues Found", str(len(qa.issues_found))],
        ]
        qa_table = Table(qa_summary, colWidths=[4.5 * cm, 4.5 * cm, 4.5 * cm, 4.5 * cm])
        qa_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (0, -1), LIGHT_BLUE),
            ("BACKGROUND", (2, 0), (2, -1), LIGHT_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(qa_table)
        story.append(Spacer(1, 8))

        checks_header = ["Check Name", "Status", "Confidence", "Notes"]
        checks_data = [checks_header]
        for check in qa.checks:
            checks_data.append([
                check.check_name,
                "PASS" if check.passed else "FAIL",
                f"{check.confidence:.0f}%",
                check.note or "",
            ])
        c_table = Table(checks_data, colWidths=[6 * cm, 2 * cm, 2.5 * cm, 7 * cm])
        c_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
            ("PADDING", (0, 0), (-1, -1), 3),
        ]))
        # Color PASS/FAIL cells
        for i, check in enumerate(qa.checks, start=1):
            color = GREEN if check.passed else RED
            c_table.setStyle(TableStyle([("TEXTCOLOR", (1, i), (1, i), color)]))
        story.append(c_table)

    # Footer
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
    story.append(Paragraph(
        f"Generated by AI Bank Statement Analyser | {datetime.now().strftime('%d %B %Y %H:%M')}",
        ParagraphStyle("Footer", fontSize=7, textColor=GRAY, alignment=TA_CENTER),
    ))

    doc.build(story)
    return output_path


def generate_excel_report(analysis: AnalysisResponse, output_path: str) -> str:
    """Generate Excel workbook with 3 sheets: Transactions, Monthly Summary, Risk Report."""
    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1a2e4a")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="F5F6FA")
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    def style_data_row(ws, row_num, col_count, alt=False):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            if alt:
                cell.fill = ALT_FILL
            cell.border = BORDER
            cell.font = Font(size=9)

    # Sheet 1: Transactions
    ws_txn = wb.active
    ws_txn.title = "Transactions"
    headers = ["Date", "Narration", "Debit (₹)", "Credit (₹)", "Balance (₹)", "Category", "Type", "Salary", "EMI", "Bounce", "Recurring", "Suspicious", "Confidence"]
    ws_txn.append(headers)
    style_header_row(ws_txn, 1, len(headers))

    if analysis.transactions:
        for i, t in enumerate(analysis.transactions, start=2):
            ws_txn.append([
                t.date, t.narration,
                t.debit or "", t.credit or "", t.balance or "",
                t.category.value if t.category else "",
                t.transaction_type or "",
                "Yes" if t.is_salary else "", "Yes" if t.is_emi else "",
                "Yes" if t.is_bounce else "", "Yes" if t.is_recurring else "",
                "Yes" if t.is_suspicious else "",
                f"{t.confidence*100:.0f}%",
            ])
            style_data_row(ws_txn, i, len(headers), alt=(i % 2 == 0))

    col_widths = [12, 45, 14, 14, 14, 22, 10, 8, 8, 8, 10, 10, 11]
    for i, width in enumerate(col_widths, start=1):
        ws_txn.column_dimensions[get_column_letter(i)].width = width

    # Sheet 2: Monthly Summary
    ws_month = wb.create_sheet("Monthly Summary")
    m_headers = ["Month", "Total Credits", "Total Debits", "Net Cash Flow", "Avg Balance", "Min Balance", "Max Balance", "Transactions", "Salary", "EMI", "Bounces"]
    ws_month.append(m_headers)
    style_header_row(ws_month, 1, len(m_headers))

    if analysis.analytics and analysis.analytics.monthly_stats:
        for i, ms in enumerate(analysis.analytics.monthly_stats, start=2):
            ws_month.append([
                ms.month, ms.total_credits, ms.total_debits, ms.net_cash_flow,
                ms.average_balance, ms.min_balance, ms.max_balance,
                ms.transaction_count, ms.salary_credits, ms.emi_debits, ms.bounce_count,
            ])
            style_data_row(ws_month, i, len(m_headers), alt=(i % 2 == 0))
    for i in range(1, len(m_headers) + 1):
        ws_month.column_dimensions[get_column_letter(i)].width = 16

    # Sheet 3: Risk Report
    ws_risk = wb.create_sheet("Risk Report")
    if analysis.analytics:
        cw = analysis.analytics.creditworthiness
        risk_rows = [
            ["BSA Score", cw.bsa_score],
            ["Risk Category", cw.risk_category],
            ["FOIR (%)", cw.foir or "N/A"],
            ["Income Stability Index", cw.income_stability_index],
            ["Avg Monthly Income", cw.average_monthly_income],
            ["Avg Monthly EMI", cw.average_monthly_emi],
            ["Disposable Income", cw.disposable_income],
            ["Max Eligible EMI", cw.max_eligible_emi],
        ]
        ws_risk.append(["Metric", "Value"])
        style_header_row(ws_risk, 1, 2)
        for i, row in enumerate(risk_rows, start=2):
            ws_risk.append(row)
            style_data_row(ws_risk, i, 2, alt=(i % 2 == 0))

        ws_risk.append([])
        ws_risk.append(["Risk Flags"])
        ws_risk.append(["Flag Type", "Severity", "Description"])
        flag_row = ws_risk.max_row
        style_header_row(ws_risk, flag_row, 3)
        for flag in cw.risk_flags:
            ws_risk.append([flag.flag_type, flag.severity.upper(), flag.description])

        ws_risk.column_dimensions["A"].width = 28
        ws_risk.column_dimensions["B"].width = 18
        ws_risk.column_dimensions["C"].width = 50

    wb.save(output_path)
    return output_path
